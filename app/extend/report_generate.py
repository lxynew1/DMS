import json
import os
from uuid import uuid1

from flask import request,send_from_directory
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, colors

from . import extend
import zipfile

dict = {1: 31, 3: 31, 4: 30, 5: 31, 6: 30, 7: 31, 8: 31, 9: 30, 10: 31, 11: 30, 12: 31}


def maxDate(year, month):
    year = int(year)
    month = int(month)
    if (year % 100 == 0 and year % 4 == 0) or (year % 4 != 0):
        dict[2] = 28
    else:
        dict[2] = 29
    return dict[month]


def allDay(year):
    print(year)
    all_day_list = []
    year = int(year)
    for i in range(12):
        max_date = maxDate(year, i + 1)
        for j in range(max_date):
            year_month_day = '{0}-{1}-{2}'.format(year, str(i + 1).zfill(2), str(j + 1).zfill(2))
            all_day_list.append(year_month_day)
    return all_day_list


def reportGennrate(system, name, date, folder):
    # 配置当前工作目录
    os.chdir(r'C:\Users\black\Desktop\信息系统日常巡检记录')
    # 定义处理模板
    template = r'信息系统日常巡检记录（智慧住建专项应用）-公租局财务核算系统.xlsx'

    # 定义日期字典
    date_dict = {
        "1": "31",
        "2": "28",
        "3": "31",
        "4": "30",
        "5": "31",
        "6": "30",
        "7": "31",
        "8": "31",
        "9": "30",
        "10": "31",
        "11": "30",
        "12": "31",
    }

    #定义各个需要的日期格式
    year_month_day = date.split('-')
    datestr = "{0}月{1}日".format(year_month_day[1],year_month_day[2])
    datefilename = date.replace('-','')

    #创建一个文件夹
    if_exist = os.path.exists(folder)
    if not if_exist:
        os.makedirs(folder)

    # 操作excel
    wb = load_workbook(filename=template)
    sheet_ranges = wb['巡检记录']
    # 修改日期
    sheet_ranges.cell(2, 1).value = "信息系统名称：" + system
    sheet_ranges.cell(5, 1).value = datestr
    sheet_ranges.cell(5, 8).value = name

    border_set = Border(left=Side(style='thin', color=colors.BLACK),
                        right=Side(style='thin', color=colors.BLACK),
                        top=Side(style='thin', color=colors.BLACK),
                        bottom=Side(style='thin', color=colors.BLACK))
    for i in range(13):
        for j in range(9):
            sheet_ranges.cell(i + 1, j + 1).border = border_set

    # 待压索列表
    file_list = []
    newfile = r'{0}/信息系统日常巡检记录（智慧住建专项应用）-{1}{2}.xlsx'.format(folder,system,datefilename)
    file_list.append(newfile)
    wb.save(newfile)

    # 把上述文件移动到一个文件夹中并且压缩
    start_dir = folder  # 要压缩的文件夹路径
    file_news = folder + '.zip'  # 压缩后文件夹的名字

    z = zipfile.ZipFile(file_news, 'w', zipfile.ZIP_DEFLATED)
    for dir_path, dir_names, file_names in os.walk(start_dir):
        f_path = dir_path.replace(start_dir, '')  # 这一句很重要，不replace的话，就从根目录开始复制
        f_path = f_path and f_path + os.sep or ''  # 实现当前文件夹以及包含的所有文件的压缩
        for filename in file_names:
            z.write(os.path.join(dir_path, filename), f_path + filename)
    z.close()
    return r'C:\Users\black\Desktop\信息系统日常巡检记录\{0}.zip'.format(folder)



@extend.route('/report', methods=["GET"])
def reportGennrateApi():
    standard_respnonse = {
        "err_code": 200,
        "err_info": "",
        "data": "",
    }
    # 接收参数
    system = request.values.get('system')
    name = request.values.get('name')
    year = request.values.get('year')
    # 生成一个文件夹的随机值
    fodler = str(uuid1())
    all_day = allDay(year=year)
    for date in all_day:
        reportGennrate(system=system,
                       name=name,
                       date=date,
                       folder=fodler)
    return send_from_directory(r'C:\Users\black\Desktop\信息系统日常巡检记录',r'{0}.zip'.format(fodler),as_attachment=True)
    # return json.dumps(standard_respnonse, ensure_ascii=False)  # ensure_ascii=False才能输出中文
